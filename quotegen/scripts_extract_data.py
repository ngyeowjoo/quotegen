"""
Extract master data from the source Excel Quote Generator into clean JSON files.
Run once (and whenever the source Excel is updated) to refresh data/*.json.

Usage:
    python scripts_extract_data.py /path/to/Quote_Generator.xlsx
"""
import sys
import json
import openpyxl
from pathlib import Path

DATA_DIR = Path(__file__).parent / "data"
DATA_DIR.mkdir(exist_ok=True)


def load(path):
    return openpyxl.load_workbook(path, data_only=True)


def extract_drivers(wb):
    """Country/city master data: hours, shrinkage, attrition, overheads, HOOPs."""
    ws = wb["Drivers"]
    headers = [c.value for c in ws[3]]
    rows = []
    for row in ws.iter_rows(min_row=4, max_row=40):
        vals = [c.value for c in row]
        if not vals[0]:
            continue
        rec = dict(zip(headers, vals))
        rows.append({
            "biz_unit": rec.get("Biz unit"),
            "country": rec.get("Country"),
            "city": rec.get("City"),
            "paid_hrs": rec.get("Paid hrs"),
            "logged_hrs": rec.get("Logged hrs"),
            "productive_hrs": rec.get("Productive hrs"),
            "log_shrinkage": rec.get("Log shrinkage"),
            "prod_shrinkage": rec.get("Prod shrinkage"),
            "monthly_attrition": rec.get("Monthly Attrition"),
            "annual_wage_inflation": rec.get("Annual Wage inflation"),
            "staff_tenure_months": rec.get("Staff tenure (months)"),
            "social_security_pct": rec.get("Social security %"),
            "night_premium": rec.get("Night Premium"),
            "night_allowance_lc": rec.get("Night Allowance (LC$)"),
            "shared_services_lc": rec.get("Shared Services - HR, Finance,IT,Mgmt"),
            "it_telecom_lc": rec.get("IT & Telecom Costs"),
            "facilities_lc": rec.get("Facilities / Establishment"),
            "general_overheads_lc": rec.get("General Overheads"),
            "total_overheads_lc": rec.get("Total Overheads"),
            "hoops": rec.get("HOOPs (Hours of operation)"),
        })
    return rows


def extract_salary_matrix(wb):
    """SalMat: role x campaign type salary + premium build-up (computed values)."""
    ws = wb["SalMat"]
    rows = []
    for row in ws.iter_rows(min_row=9, max_row=148):
        b, c = row[1].value, row[2].value
        if not b or not c:
            continue
        rows.append({
            "campaign_type": b,
            "role": c,
            "key": row[3].value,
            "native_flag": row[4].value,
            "base_salary": row[6].value,
            "salary_adjustment": row[7].value,
            "language_native_premium": row[8].value,
            "complexity_premium": row[9].value,
            "tenure_premium": row[10].value,
            "other_premium": row[11].value,
            "total_salaries": row[12].value,
            "bonus_annual": row[13].value,
            "incentive": row[14].value,
            "cash_benefit_1": row[15].value,
            "cash_benefit_2": row[16].value,
            "total_bonus_benefits": row[17].value,
        })
    return rows


def extract_overheads(wb):
    """OH tab: per-FTE overhead cost by mode (blocks per LOB - here LOB1 block)."""
    ws = wb["OH"]
    modes = ["TDCX_INCL_IT", "TDCX_EXCL_IT", "REMOTE", "CLIENT_SITE", "HYBRID"]
    cols = [4, 6, 8, 10, 12]  # D, F, H, J, L
    buckets = ["shared_services", "it_telecom", "facilities", "general_overheads", "total_overheads"]
    bucket_rows = [7, 8, 9, 10, 11]
    out = []
    for mode, col in zip(modes, cols):
        rec = {"overhead_mode": mode}
        for bucket, r in zip(buckets, bucket_rows):
            rec[bucket] = ws.cell(row=r, column=col).value
        out.append(rec)
    return out


def extract_fx(wb):
    ws = wb["FX"]
    currencies = []
    for row in ws.iter_rows(min_row=5, max_row=30):
        code = row[1].value
        rate_to_usd = row[3].value  # column D = rate vs USD (per USD)
        if code and rate_to_usd is not None:
            currencies.append({"currency": code, "rate_per_usd": rate_to_usd})
    return currencies


def extract_tax(wb):
    ws = wb["VAT,SS"]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=36):
        country, biz_unit, currency, vat, ss = [c.value for c in row[:5]]
        if country:
            rows.append({
                "country": country,
                "biz_unit": biz_unit,
                "currency": currency,
                "vat_pct": vat,
                "social_security_pct": ss,
            })
    return rows


def extract_base_salary_db(wb):
    """
    DB-BaseSal: the REAL full country x role base salary matrix (city columns from F onward).
    SalMat only reflects the currently-selected country in the live workbook, so this
    is the source we need for a multi-country app.
    """
    ws = wb["DB-BaseSal"]
    city_cols = []
    for col in range(6, 70):
        city = ws.cell(row=2, column=col).value
        if city:
            city_cols.append((col, city))

    rows = []
    for row in ws.iter_rows(min_row=4, max_row=143):
        campaign_type = row[1].value
        seniority = row[2].value
        native_flag = row[4].value
        if not campaign_type or not seniority:
            continue
        for col, city in city_cols:
            base_salary = row[col - 1].value
            if base_salary is not None:
                rows.append({
                    "campaign_type": campaign_type,
                    "role": seniority,
                    "native_flag": native_flag,
                    "city": city,
                    "base_salary": base_salary,
                })
    return rows


def extract_span_ratios(wb, sheet="SpanLOB1"):
    """Fully-loaded support add-on per agent (row 6: J=add-on, R-U = component breakdown)."""
    ws = wb[sheet]
    return {
        "fully_loaded_addon": ws["J6"].value,
        "staff_cost_component": ws["R6"].value,
        "overhead_component": ws["S6"].value,
        "penalty_component": ws["T6"].value,
        "margin_component": ws["U6"].value,
    }


def extract_campaign_types(wb):
    """
    Campaign type display names mapped to the underscore key form used as the
    join key in SalMat / MX / base_salary_db, so the UI dropdown lines up with
    the salary data. The mapping is drawn directly from the set of keys that
    actually appear in base_salary_db (source of truth), not guessed.
    """
    ws = wb["Instruct"]
    display_to_key = {
        "ACCOUNT MGT": "ACCOUNT_MGT",
        "INSIDE SALES": "INSIDE_SALES",
        "SALES LEAD GEN": "SALES_LEAD_GEN",
        "TRUST & SAFETY": "TRUST_SAFETY",
        "CUSTOMER SERVICE": "CUSTOMER_SERVICE",
        "TECH SUPPORT": "TECH_SUPPORT",
        "CONTENT MOD": "CONTENT_MOD",
        "CUSTOM": "CUSTOM",
    }
    out = []
    for row in ws.iter_rows(min_row=6, max_row=13):
        t, d = row[4].value, row[5].value
        if t and t != "TYPE":
            out.append({"type": t, "key": display_to_key.get(t, t.replace(" ", "_")), "description": d})
    return out


def main(xlsx_path):
    wb = load(xlsx_path)

    datasets = {
        "drivers.json": extract_drivers(wb),
        "salary_matrix.json": extract_salary_matrix(wb),
        "base_salary_db.json": extract_base_salary_db(wb),
        "overheads.json": extract_overheads(wb),
        "fx_rates.json": extract_fx(wb),
        "tax_rates.json": extract_tax(wb),
        "span_ratios_lob1.json": extract_span_ratios(wb, "SpanLOB1"),
        "campaign_types.json": extract_campaign_types(wb),
    }

    for filename, data in datasets.items():
        out_path = DATA_DIR / filename
        with open(out_path, "w") as f:
            json.dump(data, f, indent=2, default=str)
        n = len(data) if isinstance(data, list) else 1
        print(f"Wrote {out_path} ({n} records)")


if __name__ == "__main__":
    xlsx = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/uploads/Quote_Generator.xlsx"
    main(xlsx)
